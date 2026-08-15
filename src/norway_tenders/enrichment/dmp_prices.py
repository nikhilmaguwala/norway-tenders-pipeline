from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from norway_tenders.matching.matcher import parse_norwegian_number
from norway_tenders.normalise.display import normalize_strength_display

DMP_SOURCE_URL = "https://www.dmp.no/offentlig-finansiering/pris-pa-legemidler/maksimalpris"
DMP_PUBLISHER = "Direktoratet for medisinske produkter"

# maxPrice = official maximum AIP in NOK per listed pack (not AUP, GIP, turnover, or tender totals).
MAX_PRICE_DEFINITION = (
  "maxPrice is the official maximum apotekinnkjøpspris (AIP) in NOK per pack. "
  "Maximum AUP, TILBUDT GIP, discount prices, turnover, and umbrella tender values are excluded."
)

JOIN_OUTCOMES = frozenset({
  "exact_validated",
  "exact_item_conflicting_attributes",
  "exact_item_missing_validation_fields",
  "no_item_match",
  "ambiguous_item_match",
})


@dataclass(frozen=True)
class DmpPackPrice:
  varenummer: str
  product_name: str
  strength: str
  pack_description: str
  atc_code: str
  maks_aip: float | None
  effective_date: str


@dataclass
class DmpPriceIndex:
  local_file: str
  filename: str
  sha256: str
  effective_date: str
  sheet_name: str
  header_row: int
  by_varenummer: dict[str, DmpPackPrice]
  stripped_to_varenummer: dict[str, list[str]]


@dataclass
class DmpJoinResult:
  join_outcome: str
  conflict_reason: str
  dmp_row: DmpPackPrice | None
  dmp_matched_item: str
  tender_document_max_aip: float | None
  dmp_max_aip: float | None
  selected_max_price: float | None
  max_price_source: str
  max_price_effective_date: str
  price_difference: float | None
  price_difference_pct: float | None


def sha256_file(path: Path) -> str:
  digest = hashlib.sha256()
  with path.open("rb") as fh:
    for chunk in iter(lambda: fh.read(65536), b""):
      digest.update(chunk)
  return digest.hexdigest()


def _parse_effective_date(workbook_path: Path, sheet) -> str:
  rows = list(sheet.iter_rows(min_row=1, max_row=3, values_only=True))
  for row in rows:
    for cell in row:
      if cell is None:
        continue
      text = str(cell).strip()
      match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", text)
      if match:
        d, m, y = match.groups()
        return f"{y}-{m}-{d}"
  folder = workbook_path.parent.name
  match = re.search(r"(\d{4}-\d{2}-\d{2})", folder)
  if match:
    return match.group(1)
  match = re.search(r"(\d{4}-\d{2}-\d{2})", workbook_path.name)
  if match:
    return match.group(1)
  return ""


def discover_dmp_workbook(seeds_root: Path) -> Path:
  candidates = sorted(seeds_root.glob("DMP_Maximum_Prices__*/**/*.xlsx"))
  if not candidates:
    raise FileNotFoundError("No DMP maximum price workbook under data/seeds/DMP_Maximum_Prices__*/")
  return candidates[0]


def load_dmp_price_index(workbook_path: Path, *, seeds_root: Path) -> DmpPriceIndex:
  if workbook_path.read_bytes()[:2] != b"PK":
    raise ValueError(f"Not a valid XLSX/OpenXML file: {workbook_path}")

  wb = load_workbook(workbook_path, read_only=True, data_only=True)
  sheet_name = wb.sheetnames[0]
  ws = wb[sheet_name]
  effective_date = _parse_effective_date(workbook_path, ws)

  headers = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
  header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}

  by_varenummer: dict[str, DmpPackPrice] = {}
  stripped_to_varenummer: dict[str, list[str]] = {}

  for row in ws.iter_rows(min_row=4, values_only=True):
    raw_vn = row[0]
    if raw_vn is None or str(raw_vn).strip() == "":
      continue
    varenummer = str(raw_vn).strip()
    if varenummer.endswith(".0"):
      varenummer = varenummer[:-2]
    strength_raw = str(row[header_map.get("Styrke", 6)] or "").strip()
    pack_parts = [
      str(row[idx] or "").strip()
      for idx in (
        header_map.get("Pakningstype", 7),
        header_map.get("Antall beholdere", 9),
        header_map.get("Mengde per beholder", 10),
        header_map.get("Måle-enhet", 11),
      )
      if idx is not None
    ]
    pack_description = " ".join(p for p in pack_parts if p)
    atc = str(row[header_map.get("ATC-kode (pakning)", 21)] or "").strip().upper()
    maks_aip = parse_norwegian_number(row[header_map.get("Maks AIP Gyldig", 12)])

    record = DmpPackPrice(
      varenummer=varenummer,
      product_name=str(row[header_map.get("Handelsnavn", 2)] or "").strip(),
      strength=strength_raw,
      pack_description=pack_description,
      atc_code=atc,
      maks_aip=maks_aip,
      effective_date=effective_date,
    )
    by_varenummer[varenummer] = record
    stripped = varenummer.lstrip("0") or "0"
    stripped_to_varenummer.setdefault(stripped, []).append(varenummer)

  wb.close()
  rel = str(workbook_path.relative_to(seeds_root))
  return DmpPriceIndex(
    local_file=rel,
    filename=workbook_path.name,
    sha256=sha256_file(workbook_path),
    effective_date=effective_date,
    sheet_name=sheet_name,
    header_row=3,
    by_varenummer=by_varenummer,
    stripped_to_varenummer=stripped_to_varenummer,
  )


def _lookup_dmp_row(index: DmpPriceIndex, item_number: str) -> tuple[DmpPackPrice | None, str, str]:
  item = (item_number or "").strip()
  if not item:
    return None, "", "missing_item_number"

  if item in index.by_varenummer:
    return index.by_varenummer[item], item, "exact_string"

  stripped = item.lstrip("0") or "0"
  matches = index.stripped_to_varenummer.get(stripped, [])
  if len(matches) == 1:
    vn = matches[0]
    return index.by_varenummer[vn], vn, "leading_zero_canonical"
  if len(matches) > 1:
    return None, "", "ambiguous_stripped_match"
  return None, "", "no_match"


def _normalize_atc(value: str) -> str:
  return re.sub(r"\s+", "", (value or "").upper())


def _strength_compatible(source_strength: str, dmp_strength: str) -> bool:
  a = normalize_strength_display(source_strength).casefold().replace(" ", "")
  b = normalize_strength_display(dmp_strength).casefold().replace(" ", "").replace(",", ".")
  if not a or not b:
    return True
  return a in b or b in a or a == b


def _product_compatible(source_name: str, dmp_name: str) -> bool:
  if not source_name or not dmp_name:
    return True
  source_tokens = {t for t in re.split(r"\W+", source_name.casefold()) if len(t) > 2}
  dmp_tokens = {t for t in re.split(r"\W+", dmp_name.casefold()) if len(t) > 2}
  if not source_tokens or not dmp_tokens:
    return True
  return bool(source_tokens & dmp_tokens)


def join_pack_to_dmp(
  *,
  index: DmpPriceIndex,
  item_number: str,
  product_molecule: str,
  source_atc: str,
  product_name: str,
  strength: str,
  pack_size: Any,
  tender_document_max_aip: float | None,
) -> DmpJoinResult:
  dmp_row, matched_item, lookup_method = _lookup_dmp_row(index, item_number)

  if lookup_method == "ambiguous_stripped_match":
    return DmpJoinResult(
      join_outcome="ambiguous_item_match",
      conflict_reason="Multiple DMP varenummer share stripped item number",
      dmp_row=None,
      dmp_matched_item="",
      tender_document_max_aip=tender_document_max_aip,
      dmp_max_aip=None,
      selected_max_price=tender_document_max_aip,
      max_price_source="tender_document" if tender_document_max_aip is not None else "",
      max_price_effective_date="",
      price_difference=None,
      price_difference_pct=None,
    )

  if dmp_row is None:
    return DmpJoinResult(
      join_outcome="no_item_match",
      conflict_reason="",
      dmp_row=None,
      dmp_matched_item="",
      tender_document_max_aip=tender_document_max_aip,
      dmp_max_aip=None,
      selected_max_price=tender_document_max_aip,
      max_price_source="tender_document" if tender_document_max_aip is not None else "",
      max_price_effective_date="",
      price_difference=None,
      price_difference_pct=None,
    )

  conflicts: list[str] = []
  if source_atc and dmp_row.atc_code and _normalize_atc(source_atc) != _normalize_atc(dmp_row.atc_code):
    conflicts.append(f"ATC {source_atc} vs DMP {dmp_row.atc_code}")
  if not _strength_compatible(strength, dmp_row.strength):
    conflicts.append(f"strength {strength} vs DMP {dmp_row.strength}")
  if not _product_compatible(product_name, dmp_row.product_name):
    conflicts.append(f"product {product_name} vs DMP {dmp_row.product_name}")

  missing_validation = not source_atc and not strength
  if conflicts:
    outcome = "exact_item_conflicting_attributes"
  elif missing_validation:
    outcome = "exact_item_missing_validation_fields"
  else:
    outcome = "exact_validated"

  dmp_max = dmp_row.maks_aip
  selected = tender_document_max_aip
  source = ""
  effective = ""

  if tender_document_max_aip is not None:
    selected = tender_document_max_aip
    source = "tender_document"
    effective = ""
  elif outcome == "exact_validated" and dmp_max is not None:
    selected = dmp_max
    source = "dmp_current_reference"
    effective = index.effective_date

  price_diff = None
  price_diff_pct = None
  if tender_document_max_aip is not None and dmp_max is not None:
    price_diff = round(dmp_max - tender_document_max_aip, 2)
    if tender_document_max_aip:
      price_diff_pct = round(100 * price_diff / tender_document_max_aip, 2)

  if outcome != "exact_validated" and tender_document_max_aip is None:
    selected = None
    source = ""

  return DmpJoinResult(
    join_outcome=outcome,
    conflict_reason="; ".join(conflicts),
    dmp_row=dmp_row,
    dmp_matched_item=matched_item,
    tender_document_max_aip=tender_document_max_aip,
    dmp_max_aip=dmp_max,
    selected_max_price=selected,
    max_price_source=source,
    max_price_effective_date=effective,
    price_difference=price_diff,
    price_difference_pct=price_diff_pct,
  )
