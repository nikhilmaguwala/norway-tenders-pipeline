from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from norway_tenders.parsers.lis_excel import HEADER_ALIASES, _is_template_row, _map_headers, _normalise_header
from norway_tenders.parsers.pdf_parser import extract_pdf_text
from norway_tenders.validation.search_terms import (
    SearchTerm,
    build_search_terms,
    sheet_layout_hint,
    term_in_text,
)

PDF_FIELDS = [
    "buyer",
    "tender_reference",
    "notice_type",
    "publication_date",
    "deadline_date",
    "contract_start",
    "contract_end",
    "options_extensions",
    "estimated_value",
    "awarded_value",
    "currency",
    "supplier_winner",
    "price_weighting",
    "award_criteria",
    "minimum_discount",
    "maximum_aip_aup",
    "gip",
    "lot_delleveranse",
]

PDF_FIELD_PATTERNS: dict[str, list[str]] = {
    "buyer": [r"oppdragsgiver", r"contracting authority", r"sykehusinnkjøp"],
    "tender_reference": [r"anskaffelsesnr", r"referanse", r"\bLIS\s*\d+", r"\b\d{4}/\d+\b", r"2507gj-1", r"2632a"],
    "notice_type": [r"konkurranse", r"rammeavtale", r"intensjonskunngj", r"procedure"],
    "publication_date": [r"publisert", r"publication date", r"\d{2}\.\d{2}\.\d{4}"],
    "deadline_date": [r"frist", r"deadline", r"tilbudsfrist"],
    "contract_start": [r"avtalestart", r"kontraktsstart", r"startdato"],
    "contract_end": [r"avtaleslutt", r"kontraktslutt", r"sluttdato"],
    "options_extensions": [r"opsjon", r"forlengelse", r"extension"],
    "estimated_value": [r"anslått verdi", r"estimated value", r"estimert"],
    "awarded_value": [r"tildelt", r"award", r"kontraktsverdi", r"14[\s,.]*671[\s,.]*946"],
    "currency": [r"\bNOK\b", r"kroner"],
    "supplier_winner": [r"leverandør", r"vinner", r"nordic pill", r"tilbyder"],
    "price_weighting": [r"prisvekt", r"price weight"],
    "award_criteria": [r"tildelingskriter", r"award criteria"],
    "minimum_discount": [r"min(?:imum)?\s*rabatt", r"minimum discount"],
    "maximum_aip_aup": [r"maks(?:imal)?\s*aip", r"maks(?:imal)?\s*aup", r"max(?:imum)?\s*aip"],
    "gip": [r"\bGIP\b", r"grossistinnkjøpspris", r"tilbudt gip"],
    "lot_delleveranse": [r"delleveranse", r"\blot\b", r"delkontrakt"],
}


@dataclass
class ExcelCellMatch:
    local_file: str
    workbook: str
    sheet: str
    row: int
    column: str
    exact_matched_value: str
    matched_term: str
    match_type: str
    target_molecule: str
    surrounding_row: str
    detected_headers: str
    layout_classification: str


@dataclass
class PdfFieldMatch:
    local_file: str
    page: int
    field: str
    exact_matched_value: str
    matched_term: str
    match_type: str
    target_molecule: str
    evidence_snippet: str
    multi_molecule_warning: str


@dataclass
class SheetInventory:
    local_file: str
    workbook: str
    sheet: str
    used_rows: int
    used_cols: int
    header_row: int | None
    detected_headers: str
    data_row_count: int
    layout_classification: str


def _row_text(row: tuple[Any, ...]) -> str:
    return " | ".join(str(c) for c in row if c is not None and str(c).strip())


def _count_data_rows(ws, header_row: int) -> int:
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        if _is_template_row(row):
            continue
        count += 1
    return count


def _detect_header_row(rows: list[tuple[Any, ...]]) -> tuple[int | None, str]:
    for idx, row in enumerate(rows[:30], 1):
        headers = [_normalise_header(c) for c in row]
        hits = sum(
            1 for h in headers
            if h and any(
                h in aliases or any(alias in h for alias in aliases if len(alias) > 3)
                for aliases in HEADER_ALIASES.values()
            )
        )
        if hits >= 2:
            detected = " | ".join(str(c) for c in row if c is not None)
            return idx, detected
    return None, ""


def inventory_workbook(path: Path, seeds_root: Path) -> tuple[list[SheetInventory], list[ExcelCellMatch]]:
    rel = str(path.relative_to(seeds_root))
    terms = build_search_terms()
    inventories: list[SheetInventory] = []
    matches: list[ExcelCellMatch] = []

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        used_rows = len(rows)
        used_cols = max((len(r) for r in rows), default=0)
        header_row, detected_headers = _detect_header_row(rows)
        data_rows = _count_data_rows(ws, header_row or 3) if header_row else 0
        header_map = _map_headers(rows[header_row - 1]) if header_row else {}
        has_pack_cols = bool({"item_number", "product_name", "atc_code"} & set(header_map))
        layout = sheet_layout_hint(sheet_name, detected_headers, has_pack_cols)
        inventories.append(
            SheetInventory(
                local_file=rel,
                workbook=path.name,
                sheet=sheet_name,
                used_rows=used_rows,
                used_cols=used_cols,
                header_row=header_row,
                detected_headers=detected_headers,
                data_row_count=data_rows,
                layout_classification=layout,
            )
        )

        for row_idx, row in enumerate(rows, 1):
            row_text = _row_text(row)
            for col_idx, cell in enumerate(row, 1):
                if cell is None:
                    continue
                cell_text = str(cell).strip()
                if not cell_text:
                    continue
                for term in terms:
                    if not term_in_text(term.term, cell_text):
                        continue
                    matches.append(
                        ExcelCellMatch(
                            local_file=rel,
                            workbook=path.name,
                            sheet=sheet_name,
                            row=row_idx,
                            column=get_column_letter(col_idx),
                            exact_matched_value=cell_text[:500],
                            matched_term=term.term,
                            match_type=term.match_type,
                            target_molecule=term.molecule,
                            surrounding_row=row_text[:500],
                            detected_headers=detected_headers,
                            layout_classification=layout,
                        )
                    )
    wb.close()
    return inventories, matches


def inventory_pdf(path: Path, seeds_root: Path, folder_molecule: str) -> list[PdfFieldMatch]:
    rel = str(path.relative_to(seeds_root))
    text = extract_pdf_text(path) or ""
    pages = re.split(r"\f", text) if "\f" in text else [text]
    if len(pages) == 1:
        pages = re.split(r"(?=\n\s*\d+\s*\n)", text) or [text]

    terms = build_search_terms()
    results: list[PdfFieldMatch] = []

    for page_num, page_text in enumerate(pages, 1):
        snippet_base = " ".join(page_text.split())[:300]
        for term in terms:
            if term_in_text(term.term, page_text):
                multi_warn = ""
                if re.search(r"14[\s,.]*671[\s,.]*946", page_text) and term.molecule == "Paliperidone":
                    multi_warn = "Notice total NOK 14,671,946 covers multiple medicines; do not allocate to Paliperidone alone"
                results.append(
                    PdfFieldMatch(
                        local_file=rel,
                        page=page_num,
                        field="molecule_term",
                        exact_matched_value=term.term,
                        matched_term=term.term,
                        match_type=term.match_type,
                        target_molecule=term.molecule,
                        evidence_snippet=snippet_base,
                        multi_molecule_warning=multi_warn,
                    )
                )

        for field, patterns in PDF_FIELD_PATTERNS.items():
            for pattern in patterns:
                match = re.search(pattern, page_text, re.IGNORECASE)
                if not match:
                    continue
                start = max(0, match.start() - 80)
                end = min(len(page_text), match.end() + 120)
                snippet = " ".join(page_text[start:end].split())
                multi_warn = ""
                if field == "awarded_value" and re.search(r"14[\s,.]*671[\s,.]*946", page_text):
                    multi_warn = "Multi-molecule notice value; not molecule-specific"
                results.append(
                    PdfFieldMatch(
                        local_file=rel,
                        page=page_num,
                        field=field,
                        exact_matched_value=match.group(0),
                        matched_term=pattern,
                        match_type="metadata",
                        target_molecule=folder_molecule,
                        evidence_snippet=snippet[:400],
                        multi_molecule_warning=multi_warn,
                    )
                )
                break
    return results
