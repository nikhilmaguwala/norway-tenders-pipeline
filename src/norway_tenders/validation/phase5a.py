from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from norway_tenders.matching.matcher import load_molecule_config
from norway_tenders.parsers.lis_excel import HEADER_ALIASES, SheetLayout, parse_lis_prisskjema
from norway_tenders.settings import DISCOVERY_DIR, SEEDS_DIR, SOURCES_SEED
from norway_tenders.validation.file_probe import FileProbeResult, probe_file
from norway_tenders.validation.inventory import (
    ExcelCellMatch,
    PdfFieldMatch,
    SheetInventory,
    inventory_pdf,
    inventory_workbook,
)
from norway_tenders.validation.seed_config import (
    HUMAN_READABLE_SEED_FOLDERS,
    PALIPERIDONE_NOTICE_TOTAL_NOK,
    SEED_FOLDER_META,
)

LOCAL_FILE_VALIDATION_CSV = DISCOVERY_DIR / "local_file_validation.csv"
LOCAL_DOCUMENT_INVENTORY_CSV = DISCOVERY_DIR / "local_document_inventory.csv"
MOLECULE_DOCUMENT_MATCHES_CSV = DISCOVERY_DIR / "molecule_document_matches.csv"
PARSER_LAYOUT_REPORT_CSV = DISCOVERY_DIR / "parser_layout_report.csv"
PHASE5A_SUMMARY_JSON = DISCOVERY_DIR / "phase5a_summary.json"

REQUIRED_PARSER_COLUMNS = ("item_number", "product_name", "atc_code")
USEFUL_PARSER_COLUMNS = ("strength", "pack_size", "supplier", "max_price", "offered_gip", "packs_year", "active_substance")

SOURCES_EXTENDED_COLUMNS = [
    "tenderRef",
    "noticeId",
    "filename",
    "url",
    "source",
    "notes",
    "targetMolecule",
    "localFile",
    "sourceType",
    "accessStatus",
    "noticeUrl",
    "landingPage",
    "linkageNeedsReview",
    "procurementFamily",
    "sha256",
]


@dataclass
class Phase5aResult:
    validation_path: Path
    inventory_path: Path
    matches_path: Path
    parser_path: Path
    summary_path: Path
    sources_path: Path
    file_count: int
    test_note: str = ""


def _infer_source_type(filename: str) -> str:
    name = filename.casefold()
    if "prisskjema" in name or "bilag 2" in name:
        return "price_schedule"
    if "kravspesifikasjon" in name or "kravspes" in name or "bilag 3" in name:
        return "requirements_specification"
    if "konkurransebestemmelser" in name:
        return "competition_terms"
    if "rammeavtale" in name or "bilag 4" in name:
        return "framework_agreement"
    return "other"


def _folder_for_path(rel: str) -> str | None:
    top = rel.split("/", 1)[0]
    return top if top in SEED_FOLDER_META else None


def discover_seed_files(seeds_root: Path = SEEDS_DIR) -> list[Path]:
    files: list[Path] = []
    for folder_name in HUMAN_READABLE_SEED_FOLDERS:
        folder = seeds_root / folder_name
        if not folder.is_dir():
            continue
        for path in sorted(folder.rglob("*")):
            if path.is_file():
                files.append(path)
    return files


def validate_local_files(seeds_root: Path = SEEDS_DIR) -> list[FileProbeResult]:
    return [probe_file(path, seeds_root) for path in discover_seed_files(seeds_root)]


def _read_existing_sources(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    rows: list[dict[str, str]] = []
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append({col: row.get(col, "") or "" for col in reader.fieldnames or []})
    return rows


def _normalize_source_row(row: dict[str, str]) -> dict[str, str]:
    return {col: row.get(col, "") for col in SOURCES_EXTENDED_COLUMNS}


def update_sources_csv(
    validations: list[FileProbeResult],
    seeds_root: Path = SEEDS_DIR,
    path: Path = SOURCES_SEED,
) -> list[dict[str, str]]:
    existing = _read_existing_sources(path)
    normalized_existing: list[dict[str, str]] = []
    seen_local: set[str] = set()
    for row in existing:
        nr = _normalize_source_row(row)
        normalized_existing.append(nr)
        if nr.get("localFile"):
            seen_local.add(nr["localFile"])

    new_rows: list[dict[str, str]] = []
    for probe in validations:
        if probe.local_file in seen_local:
            continue
        folder_key = _folder_for_path(probe.local_file)
        if not folder_key:
            continue
        meta = SEED_FOLDER_META[folder_key]
        new_rows.append(
            {
                "tenderRef": meta.tender_ref,
                "noticeId": meta.notice_id,
                "filename": probe.filename,
                "url": "",
                "source": "local_seed",
                "notes": meta.linkage_note,
                "targetMolecule": meta.target_molecule,
                "localFile": probe.local_file,
                "sourceType": _infer_source_type(probe.filename),
                "accessStatus": meta.access_status,
                "noticeUrl": meta.notice_url,
                "landingPage": meta.landing_page,
                "linkageNeedsReview": "true" if meta.linkage_needs_review else "false",
                "procurementFamily": meta.procurement_family,
                "sha256": probe.sha256,
            }
        )

    merged = normalized_existing + new_rows
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=SOURCES_EXTENDED_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in merged:
            writer.writerow(_normalize_source_row(row))
    return merged


def _evidence_level(name: bool, atc: bool, brand_only: bool) -> str:
    if brand_only and not name and not atc:
        return "brand_only"
    if name and atc:
        return "document_name_and_atc"
    if name:
        return "document_name"
    if atc:
        return "document_atc"
    return "no_confirmation"


def _assess_parser_layout(path: Path, seeds_root: Path, target_molecule: str) -> list[dict[str, Any]]:
    if path.suffix.lower() != ".xlsx":
        return []
    rel = str(path.relative_to(seeds_root))
    reports: list[dict[str, Any]] = []

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        header_map: dict[str, int] = {}
        for idx, row in enumerate(rows[:30], 1):
            from norway_tenders.parsers.lis_excel import _map_headers

            mapping = _map_headers(row)
            if {"item_number", "product_name"} & set(mapping):
                header_row = idx
                header_map = mapping
                break

        detected = [k for k in header_map if not k.startswith("_")]
        missing = [c for c in REQUIRED_PARSER_COLUMNS if c not in detected]
        extra = [c for c in USEFUL_PARSER_COLUMNS if c in detected]
        candidate_rows = 0
        if header_row:
            from norway_tenders.parsers.lis_excel import _is_template_row

            for row in rows[header_row:]:
                if not any(c is not None and str(c).strip() for c in row):
                    continue
                if _is_template_row(row):
                    continue
                candidate_rows += 1

        parser_supported = False
        recommended = "skip_non_pack_sheet"
        if "prisskjema" in sheet_name.casefold():
            header_for_parse = header_row or 3
            try:
                packs = parse_lis_prisskjema(
                    path,
                    layout=SheetLayout(
                        sheet_name=sheet_name,
                        header_row=header_for_parse,
                        data_start_row=header_for_parse + 1,
                    ),
                )
                parser_supported = len(packs) > 0
                recommended = "use_existing_lis_parser" if parser_supported else "adapter_required"
            except Exception:
                recommended = "adapter_required"

        layout_name = "pack-bearing price schedule" if not missing else "unknown layout"
        reports.append(
            {
                "localFile": rel,
                "sheet": sheet_name,
                "headerRow": header_row or "",
                "layoutName": layout_name,
                "targetMolecule": target_molecule,
                "detectedColumns": "|".join(detected),
                "missingRequiredColumns": "|".join(missing),
                "extraUsefulColumns": "|".join(extra),
                "candidateRows": candidate_rows,
                "existingParserSupported": str(parser_supported).lower(),
                "recommendedAction": recommended,
            }
        )
    wb.close()
    return reports


def _build_molecule_matches(
    excel_matches: list[ExcelCellMatch],
    pdf_matches: list[PdfFieldMatch],
    validations: list[FileProbeResult],
    parser_reports: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    molecules = load_molecule_config()
    rows: list[dict[str, Any]] = []
    summary_by_molecule: dict[str, dict[str, Any]] = {
        m: {
            "explicit_name_confirmed": False,
            "atc_confirmed": False,
            "pack_bearing_rows_found": False,
            "candidate_pack_row_count": 0,
            "item_numbers_available": False,
            "product_names_available": False,
            "strengths_available": False,
            "pack_sizes_available": False,
            "supplier_available": False,
            "volume_available": False,
            "price_available": False,
            "exact_price_type": "",
            "parser_supported": False,
            "parser_adapter_required": False,
            "evidence_levels": set(),
            "axitinib_confirmation": {},
        }
        for m in molecules
    }

    probe_by_file = {v.local_file: v for v in validations}
    parser_by_file = {r["localFile"]: r for r in parser_reports if r.get("existingParserSupported") == "true"}

    for match in excel_matches:
        mol = match.target_molecule
        is_brand = match.match_type == "brand_discovery"
        is_name = match.match_type == "name"
        is_atc = match.match_type == "atc"
        if is_name:
            summary_by_molecule[mol]["explicit_name_confirmed"] = True
        if is_atc:
            summary_by_molecule[mol]["atc_confirmed"] = True
        ev = _evidence_level(is_name, is_atc, is_brand)
        if is_brand and not (summary_by_molecule[mol]["explicit_name_confirmed"] or summary_by_molecule[mol]["atc_confirmed"]):
            ev = "brand_only"
        summary_by_molecule[mol]["evidence_levels"].add(ev)
        is_pack = match.layout_classification == "pack-bearing price schedule"
        if is_pack:
            summary_by_molecule[mol]["pack_bearing_rows_found"] = True
        rows.append(
            {
                "targetMolecule": mol,
                "localFile": match.local_file,
                "sourceType": _infer_source_type(match.workbook),
                "sheetOrPage": match.sheet,
                "rowOrSection": str(match.row),
                "exactMatchedValue": match.exact_matched_value,
                "matchedTerm": match.matched_term,
                "matchType": match.match_type,
                "atcCode": match.matched_term if is_atc else "",
                "evidenceLevel": ev,
                "isPackBearing": str(is_pack).lower(),
                "parserSupported": str(match.local_file in parser_by_file).lower(),
                "warning": "brand_only_not_accepted" if ev == "brand_only" else "",
            }
        )

    for match in pdf_matches:
        if match.field != "molecule_term":
            continue
        mol = match.target_molecule
        is_atc = bool(re.match(r"^[A-Z]\d{2}", match.matched_term))
        is_name = not is_atc and match.match_type != "brand_discovery"
        if is_name:
            summary_by_molecule[mol]["explicit_name_confirmed"] = True
        if is_atc:
            summary_by_molecule[mol]["atc_confirmed"] = True
        ev = _evidence_level(is_name, is_atc, match.match_type == "brand_discovery")
        summary_by_molecule[mol]["evidence_levels"].add(ev)
        warning = match.multi_molecule_warning
        rows.append(
            {
                "targetMolecule": mol,
                "localFile": match.local_file,
                "sourceType": _infer_source_type(Path(match.local_file).name),
                "sheetOrPage": str(match.page),
                "rowOrSection": match.field,
                "exactMatchedValue": match.exact_matched_value,
                "matchedTerm": match.matched_term,
                "matchType": match.match_type,
                "atcCode": match.matched_term if is_atc else "",
                "evidenceLevel": ev,
                "isPackBearing": "false",
                "parserSupported": "false",
                "warning": warning,
            }
        )

    for report in parser_reports:
        mol = report["targetMolecule"]
        if report.get("existingParserSupported") == "true":
            summary_by_molecule[mol]["parser_supported"] = True
        elif report.get("recommendedAction") == "adapter_required" and "prisskjema" in report.get("sheet", "").casefold():
            summary_by_molecule[mol]["parser_adapter_required"] = True
        detected = report.get("detectedColumns", "")
        if "item_number" in detected:
            summary_by_molecule[mol]["item_numbers_available"] = True
        if "product_name" in detected:
            summary_by_molecule[mol]["product_names_available"] = True
        if "strength" in detected:
            summary_by_molecule[mol]["strengths_available"] = True
        if "pack_size" in detected:
            summary_by_molecule[mol]["pack_sizes_available"] = True
        if "supplier" in detected:
            summary_by_molecule[mol]["supplier_available"] = True
        if "packs_year" in detected or "packs_sold_col" in detected:
            summary_by_molecule[mol]["volume_available"] = True
        if "max_price" in detected:
            summary_by_molecule[mol]["price_available"] = True
            summary_by_molecule[mol]["exact_price_type"] = "max_aip"
        if "offered_gip" in detected and not summary_by_molecule[mol]["price_available"]:
            summary_by_molecule[mol]["exact_price_type"] = "offered_gip_not_max_price"
        summary_by_molecule[mol]["candidate_pack_row_count"] = max(
            summary_by_molecule[mol]["candidate_pack_row_count"],
            int(report.get("candidateRows") or 0),
        )

    ax = summary_by_molecule["Axitinib"]
    axitinib_name = any(
        m.target_molecule == "Axitinib" and m.match_type == "name" for m in excel_matches
    ) or any(
        m.target_molecule == "Axitinib" and m.field == "molecule_term" and m.matched_term in ("axitinib",)
        for m in pdf_matches
    )
    ax_l01ek01 = any(m.target_molecule == "Axitinib" and m.matched_term == "L01EK01" for m in excel_matches)
    ax_l01xe17 = any(m.target_molecule == "Axitinib" and m.matched_term == "L01XE17" for m in excel_matches)
    ax_inlyta_only = (
        any(m.target_molecule == "Axitinib" and m.match_type == "brand_discovery" for m in excel_matches)
        and not axitinib_name and not ax_l01ek01 and not ax_l01xe17
    )
    ax["axitinib_confirmation"] = {
        "by_axitinib_name": axitinib_name,
        "by_L01EK01": ax_l01ek01,
        "by_L01XE17": ax_l01xe17,
        "by_inlyta_only": ax_inlyta_only,
        "no_confirmation": not (axitinib_name or ax_l01ek01 or ax_l01xe17 or ax_inlyta_only),
        "lis_2207_confirms_via": (
            "L01EK01_and_Inlyta_brand"
            if ax_l01ek01 and not axitinib_name
            else ("axitinib_name" if axitinib_name else ("inlyta_only" if ax_inlyta_only else "no_confirmation"))
        ),
    }

    for mol, data in summary_by_molecule.items():
        data["evidence_levels"] = sorted(data["evidence_levels"])

    return rows, summary_by_molecule


def run_phase5a(seeds_root: Path = SEEDS_DIR) -> Phase5aResult:
    DISCOVERY_DIR.mkdir(parents=True, exist_ok=True)

    validations = validate_local_files(seeds_root)
    sources = update_sources_csv(validations, seeds_root)

    sheet_inventories: list[SheetInventory] = []
    excel_matches: list[ExcelCellMatch] = []
    pdf_matches: list[PdfFieldMatch] = []
    parser_reports: list[dict[str, Any]] = []

    for path in discover_seed_files(seeds_root):
        folder_key = path.relative_to(seeds_root).parts[0]
        meta = SEED_FOLDER_META.get(folder_key)
        target = meta.target_molecule if meta else ""
        if path.suffix.lower() == ".xlsx":
            inv, matches = inventory_workbook(path, seeds_root)
            sheet_inventories.extend(inv)
            excel_matches.extend(matches)
            parser_reports.extend(_assess_parser_layout(path, seeds_root, target))
        elif path.suffix.lower() == ".pdf":
            pdf_matches.extend(inventory_pdf(path, seeds_root, target))

    with LOCAL_FILE_VALIDATION_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "localFile", "filename", "extension", "detectedType", "fileSize",
                "sha256", "isValid", "hasEmbeddedText", "validationWarning",
            ],
        )
        writer.writeheader()
        for v in validations:
            writer.writerow(
                {
                    "localFile": v.local_file,
                    "filename": v.filename,
                    "extension": v.extension,
                    "detectedType": v.detected_type,
                    "fileSize": v.file_size,
                    "sha256": v.sha256,
                    "isValid": str(v.is_valid).lower(),
                    "hasEmbeddedText": str(v.has_embedded_text).lower(),
                    "validationWarning": v.validation_warning,
                }
            )

    with LOCAL_DOCUMENT_INVENTORY_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "localFile", "fileType", "workbook", "sheet", "page", "row", "column",
                "field", "exactMatchedValue", "matchedTerm", "matchType", "targetMolecule",
                "surroundingContext", "detectedHeaders", "layoutClassification", "evidenceSnippet",
                "multiMoleculeWarning",
            ],
        )
        writer.writeheader()
        for inv in sheet_inventories:
            writer.writerow(
                {
                    "localFile": inv.local_file,
                    "fileType": "xlsx",
                    "workbook": inv.workbook,
                    "sheet": inv.sheet,
                    "page": "",
                    "row": "",
                    "column": "",
                    "field": "sheet_inventory",
                    "exactMatchedValue": f"rows={inv.used_rows};cols={inv.used_cols};dataRows={inv.data_row_count}",
                    "matchedTerm": "",
                    "matchType": "",
                    "targetMolecule": SEED_FOLDER_META[inv.local_file.split("/")[0]].target_molecule,
                    "surroundingContext": "",
                    "detectedHeaders": inv.detected_headers,
                    "layoutClassification": inv.layout_classification,
                    "evidenceSnippet": "",
                    "multiMoleculeWarning": "",
                }
            )
        for match in excel_matches:
            writer.writerow(
                {
                    "localFile": match.local_file,
                    "fileType": "xlsx",
                    "workbook": match.workbook,
                    "sheet": match.sheet,
                    "page": "",
                    "row": match.row,
                    "column": match.column,
                    "field": "molecule_match",
                    "exactMatchedValue": match.exact_matched_value,
                    "matchedTerm": match.matched_term,
                    "matchType": match.match_type,
                    "targetMolecule": match.target_molecule,
                    "surroundingContext": match.surrounding_row,
                    "detectedHeaders": match.detected_headers,
                    "layoutClassification": match.layout_classification,
                    "evidenceSnippet": "",
                    "multiMoleculeWarning": "",
                }
            )
        for match in pdf_matches:
            writer.writerow(
                {
                    "localFile": match.local_file,
                    "fileType": "pdf",
                    "workbook": "",
                    "sheet": "",
                    "page": match.page,
                    "row": "",
                    "column": "",
                    "field": match.field,
                    "exactMatchedValue": match.exact_matched_value,
                    "matchedTerm": match.matched_term,
                    "matchType": "pdf_text",
                    "targetMolecule": match.target_molecule,
                    "surroundingContext": "",
                    "detectedHeaders": "",
                    "layoutClassification": "",
                    "evidenceSnippet": match.evidence_snippet,
                    "multiMoleculeWarning": match.multi_molecule_warning,
                }
            )

    molecule_rows, molecule_summary = _build_molecule_matches(
        excel_matches, pdf_matches, validations, parser_reports,
    )
    with MOLECULE_DOCUMENT_MATCHES_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "targetMolecule", "localFile", "sourceType", "sheetOrPage", "rowOrSection",
                "exactMatchedValue", "matchedTerm", "matchType", "atcCode", "evidenceLevel",
                "isPackBearing", "parserSupported", "warning",
            ],
        )
        writer.writeheader()
        writer.writerows(molecule_rows)

    with PARSER_LAYOUT_REPORT_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "localFile", "sheet", "headerRow", "layoutName", "targetMolecule",
                "detectedColumns", "missingRequiredColumns", "extraUsefulColumns",
                "candidateRows", "existingParserSupported", "recommendedAction",
            ],
        )
        writer.writeheader()
        writer.writerows(parser_reports)

    total_candidate_rows = sum(
        int(r.get("candidateRows") or 0)
        for r in parser_reports
        if r.get("layoutName") == "pack-bearing price schedule"
    )
    linkage_reviews = [
        meta.folder for meta in SEED_FOLDER_META.values() if meta.linkage_needs_review
    ]

    summary = {
        "phase": "5A",
        "files_validated": [asdict(v) for v in validations],
        "file_count": len(validations),
        "valid_count": sum(1 for v in validations if v.is_valid),
        "invalid_count": sum(1 for v in validations if not v.is_valid),
        "sources_row_count": len(sources),
        "molecule_confirmation": molecule_summary,
        "candidate_pack_rows_by_molecule": {
            m: molecule_summary[m]["candidate_pack_row_count"] for m in molecule_summary
        },
        "parser_layouts_detected": sorted({r["layoutName"] for r in parser_reports}),
        "parser_support_by_workbook": {
            local_file: "true"
            if any(
                r["localFile"] == local_file and r.get("existingParserSupported") == "true"
                for r in parser_reports
            )
            else "false"
            for local_file in {r["localFile"] for r in parser_reports}
        },
        "parser_adapters_required": sorted(
            {r["localFile"] for r in parser_reports if r.get("recommendedAction") == "adapter_required"}
        ),
        "field_availability_by_molecule": {
            m: {k: v for k, v in data.items() if k not in {"evidence_levels", "axitinib_confirmation"}}
            for m, data in molecule_summary.items()
        },
        "missing_evidence": {
            m: [
                gap for gap, present in [
                    ("name", data["explicit_name_confirmed"]),
                    ("atc", data["atc_confirmed"]),
                    ("pack_rows", data["pack_bearing_rows_found"]),
                ] if not present
            ]
            for m, data in molecule_summary.items()
        },
        "defensible_row_estimate": {
            "low": max(10, total_candidate_rows // 3),
            "high": min(120, total_candidate_rows),
            "likely_in_range_40_120": 40 <= total_candidate_rows <= 120,
        },
        "value_allocation_risks": [
            "Paliperidone notice 682047-2022 total NOK 14,671,946 is multi-molecule; do not assign to Paliperidone alone",
            "Offered GIP must never map to maxPrice",
            "LIS 2207 oncology umbrella covers many molecules; allocate only Axitinib-attributed rows",
            "Everolimus tender 2632a includes mycophenolic acid; filter by ATC/name",
        ],
        "linkage_needs_review": linkage_reviews,
        "paliperidone_safety_context": {
            "noticeId": "682047-2022",
            "preserved_notice_total_nok": PALIPERIDONE_NOTICE_TOTAL_NOK,
            "do_not_allocate_notice_total_to_molecule": True,
            "known_facts": [
                "Paliperidon", "N05AX13", "lot 90", "low-dose injection",
                "Nordic Pill AB", "award 2022-12-02", "direct award",
            ],
        },
        "axitinib_evidence": molecule_summary["Axitinib"]["axitinib_confirmation"],
    }
    PHASE5A_SUMMARY_JSON.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")

    return Phase5aResult(
        validation_path=LOCAL_FILE_VALIDATION_CSV,
        inventory_path=LOCAL_DOCUMENT_INVENTORY_CSV,
        matches_path=MOLECULE_DOCUMENT_MATCHES_CSV,
        parser_path=PARSER_LAYOUT_REPORT_CSV,
        summary_path=PHASE5A_SUMMARY_JSON,
        sources_path=SOURCES_SEED,
        file_count=len(validations),
    )
