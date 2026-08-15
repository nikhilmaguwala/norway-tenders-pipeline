from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from norway_tenders.matching.matcher import parse_norwegian_number, parse_pack_size
from norway_tenders.models import PackRecord, Provenance, TenderLevelEvidence
from norway_tenders.normalise.display import normalize_pack_volume, parse_pack_volume

logger = logging.getLogger(__name__)

# Header alias map: canonical field -> known Norwegian/English aliases
HEADER_ALIASES: dict[str, list[str]] = {
    "item_number": [
        "varenr",
        "varenummer",
        "artikkelnummer",
        "item number",
        "item no",
    ],
    "product_name": ["varenavn", "varebetegnelse", "produktnavn", "legemiddel", "product name", "preparat"],
    "active_substance": ["virkestoff", "active substance", "molecule"],
    "strength": ["styrke", "strength", "dose"],
    "pack_size": ["pakningsstørrelse", "pakning", "pack size", "kvantum", "enhet", "pakkestørrelse"],
    "atc_code": ["atc", "atc-kode", "atc code", "atc kode"],
    "supplier": ["leverandør", "supplier", "produsent", "firma"],
    "max_price": ["maks aip", "maksimal aip", "max aip", "maks pris", "max price"],
    "offered_gip": ["tilbudt gip", "gip", "tilbudt pris", "offered price"],
    "packs_year": ["pakninger", "antall pakninger", "volum", "forbruk"],
    "packs_sold_col": ["pakningssalg", "solgte pakninger siste 12 mnd"],
}


@dataclass
class SheetLayout:
    sheet_name: str = "Prisskjema"
    header_row: int = 3
    data_start_row: int = 4


def _normalise_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().casefold())


def _detect_year_column(header: str) -> int | None:
    match = re.search(r"pakninger\s*(\d{4})", header, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.search(r"(\d{4})\s*pakninger", header, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def _map_headers(row_values: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    year_cols: dict[int, int] = {}
    for idx, cell in enumerate(row_values):
        header = _normalise_header(cell)
        if not header:
            continue
        year = _detect_year_column(header)
        if year is not None:
            year_cols[year] = idx
            continue
        for field, aliases in HEADER_ALIASES.items():
            if header in aliases or any(header == alias for alias in aliases):
                mapping.setdefault(field, idx)
                break
            if field == "pack_size" and "pakningssalg" in header:
                continue
            if any(alias in header for alias in aliases if len(alias) > 3):
                mapping.setdefault(field, idx)
    if year_cols:
        mapping["_year_cols"] = year_cols  # type: ignore[assignment]
    return mapping


def _is_template_row(values: tuple[Any, ...]) -> bool:
    joined = " ".join(str(v) for v in values if v is not None).casefold()
    skip_markers = [
        "eksempel",
        "utfylles av",
        "ikke fyll ut",
        "mal",
        "template",
        "sum",
        "totalt",
    ]
    return any(marker in joined for marker in skip_markers)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_lis_prisskjema(
    path: Path,
    *,
    layout: SheetLayout | None = None,
    source_url: str = "",
) -> list[PackRecord]:
    """Parse a Sykehusinnkjøp LIS Prisskjema workbook."""
    layout = layout or SheetLayout()
    if not path.exists():
        raise FileNotFoundError(path)
    if path.read_bytes()[:2] != b"PK":
        raise ValueError(f"Not a valid xlsx file: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = layout.sheet_name if layout.sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < layout.header_row:
        return []

    header_map = _map_headers(rows[layout.header_row - 1])
    year_cols: dict[int, int] = header_map.pop("_year_cols", {})  # type: ignore[arg-type]

    packs: list[PackRecord] = []
    for row_idx, row in enumerate(rows[layout.data_start_row - 1 :], start=layout.data_start_row):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        if _is_template_row(row):
            continue

        item_number = _cell_str(row[header_map["item_number"]]) if "item_number" in header_map else ""
        product_name = _cell_str(row[header_map["product_name"]]) if "product_name" in header_map else ""
        if not item_number and not product_name:
            continue

        strength = _cell_str(row[header_map["strength"]]) if "strength" in header_map else ""
        pack_raw = row[header_map["pack_size"]] if "pack_size" in header_map else ""
        atc = _cell_str(row[header_map["atc_code"]]) if "atc_code" in header_map else ""
        substance = (
            _cell_str(row[header_map["active_substance"]])
            if "active_substance" in header_map
            else ""
        )
        supplier = _cell_str(row[header_map["supplier"]]) if "supplier" in header_map else ""

        max_price = None
        if "max_price" in header_map:
            max_price = parse_norwegian_number(row[header_map["max_price"]])

        offered_gip = None
        if "offered_gip" in header_map:
            offered_gip = parse_norwegian_number(row[header_map["offered_gip"]])

        packs_year = None
        packs_sold = None
        volume_label = ""
        volume_warning = ""
        if year_cols:
            latest_year = max(year_cols)
            packs_year = latest_year
            packs_sold, volume_warning = parse_pack_volume(
                parse_norwegian_number(row[year_cols[latest_year]])
            )
            volume_label = f"PAKNINGER {latest_year}"
        elif "packs_sold_col" in header_map:
            raw_volume = row[header_map["packs_sold_col"]]
            packs_sold, volume_warning = parse_pack_volume(raw_volume)
            volume_label = "Pakningssalg"

        provenance = Provenance(
            source_url=source_url,
            filename=path.name,
            sheet=sheet_name,
            row=row_idx,
            raw_values={
                "item_number": item_number,
                "product_name": product_name,
                "strength": strength,
                "pack_size": pack_raw,
                "atc_code": atc,
                "active_substance": substance,
                "packs_year": packs_year,
                "packs_sold": packs_sold,
                "volume_label": volume_label,
                "volume_warning": volume_warning,
                "offered_gip": offered_gip,
            },
        )

        packs.append(
            PackRecord(
                item_number=item_number,
                product_name=product_name,
                strength=strength,
                pack_size=parse_pack_size(pack_raw),
                atc_code=atc,
                supplier=supplier,
                max_price=max_price,
                offered_gip=offered_gip,
                packs_sold_last_12m=packs_sold,
                packs_year=packs_year,
                provenance=provenance,
            )
        )

    wb.close()
    return packs


def parse_kravspec_omfang(
    path: Path,
    *,
    sheet_name: str = "3. Omfang",
    source_url: str = "",
) -> dict[str, Any]:
    """Extract historical turnover evidence from Kravspesifikasjon Omfang sheet."""
    if path.read_bytes()[:2] != b"PK":
        raise ValueError(f"Not a valid xlsx file: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    target = sheet_name if sheet_name in wb.sheetnames else None
    if target is None:
        for name in wb.sheetnames:
            if "omfang" in name.casefold():
                target = name
                break
    if target is None:
        wb.close()
        return {}

    ws = wb[target]
    evidence: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if len(cells) < 2:
            continue
        text = " ".join(str(c) for c in cells)
        # Table row: ATC | Virkestoff | Omsetning i maks AIP [...]
        atc_match = re.search(r"\b([A-Z]\d{2}[A-Z]{2}\d{2})\b", text)
        turnover = None
        for cell in cells:
            val = parse_norwegian_number(cell)
            if val is not None and val > 1_000_000:
                turnover = val
                break
        if atc_match and turnover is not None:
            evidence["historical_turnover_aip"] = turnover
            evidence["atc_code"] = atc_match.group(1)
            evidence["label"] = text.strip()
            break
        if "omsetning" in text.casefold() and "aip" in text.casefold():
            for cell in cells:
                val = parse_norwegian_number(cell)
                if val is not None and val > 1_000_000:
                    evidence["historical_turnover_aip"] = val
                    evidence["label"] = text.strip()
                    break
    wb.close()
    return evidence


def parse_kravspec_product_requirements(
    path: Path,
    *,
    sheet_name: str = "4. Krav til prod.",
    source_url: str = "",
) -> TenderLevelEvidence:
    """Extract tender-level pricing rules from Kravspesifikasjon product requirements."""
    if path.read_bytes()[:2] != b"PK":
        raise ValueError(f"Not a valid xlsx file: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    target = sheet_name if sheet_name in wb.sheetnames else None
    if target is None:
        for name in wb.sheetnames:
            if "krav til prod" in name.casefold():
                target = name
                break
    if target is None:
        wb.close()
        return TenderLevelEvidence()

    ws = wb[target]
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        line = " ".join(str(c) for c in row if c is not None).strip()
        if line:
            lines.append(line)

    full_text = "\n".join(lines)
    evidence = TenderLevelEvidence(
        provenance=Provenance(source_url=source_url, filename=path.name, sheet=target),
    )

    if re.search(r"\bpris\b", full_text, re.IGNORECASE) and re.search(
        r"vekting|tildlingskriterier", full_text, re.IGNORECASE
    ):
        # LIS 2234: single price criterion with weight 1 (= 100 %)
        if re.search(r"pris\s*\|\s*1\b", full_text, re.IGNORECASE) or re.search(
            r"pris\s+1\b", full_text, re.IGNORECASE
        ):
            evidence.price_weighting_percent = 100.0

    discount = re.search(
        r"minimum\s+(\d+(?:[.,]\d+)?)\s*%?\s*rabatt",
        full_text,
        re.IGNORECASE,
    )
    if discount:
        evidence.min_discount_max_aip_percent = parse_norwegian_number(discount.group(1))

    if re.search(r"grossistens innkjøpspris\s*\(gip\)", full_text, re.IGNORECASE):
        evidence.offered_price_basis = "GIP in NOK per item number"

    if re.search(r"lik pris per mg", full_text, re.IGNORECASE):
        evidence.equal_price_per_mg_within_formulation = True

    wb.close()
    return evidence


def extract_workbook_buyer(path: Path) -> str:
    """Return buyer name if explicitly present in workbook text."""
    if path.read_bytes()[:2] != b"PK":
        return ""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True, max_row=40):
                text = " ".join(str(c) for c in row if c is not None)
                if re.search(r"sykehusinnkj[oø]p", text, re.IGNORECASE):
                    match = re.search(
                        r"(Sykehusinnkj[oø]p(?:\s+HF)?)",
                        text,
                        re.IGNORECASE,
                    )
                    if match:
                        name = match.group(1)
                        if "hf" in name.casefold():
                            return "Sykehusinnkjøp HF"
                        return name
    finally:
        wb.close()
    return ""
